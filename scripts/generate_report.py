"""
Generate Report Module
Purpose: Create multi-sheet Excel report with reconciliation data
"""

import pandas as pd
from pathlib import Path
from loguru import logger
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportGenerator:
    """Generate Excel reconciliation report"""
    
    def __init__(self, output_folder: str = "data/output"):
        """Initialize report generator"""
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)
        logger.info(f"ReportGenerator initialized: {self.output_folder}")
    
    def generate_report(self, merged_df: pd.DataFrame, report_date: str = None) -> Path:
        """
        Generate multi-sheet Excel report
        
        Sheets:
        1. Reconciled Data (full detail)
        2. Exceptions Only (filtered)
        3. Vendor Summary (aggregated)
        4. YTD Summary (tie-back)
        """
        logger.info("="*80)
        logger.info("GENERATING EXCEL REPORT")
        logger.info("="*80)
        
        if report_date is None:
            report_date = datetime.now().strftime("%m.%d.%Y")
        
        filename = f"Subcontractor_Accrual_Reconciliation_{report_date}.xlsx"
        filepath = self.output_folder / filename
        
        logger.info(f"Report filename: {filename}")
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            
            # Sheet 1: Reconciled Data (Full Detail)
            logger.info("Writing Sheet 1: Reconciled Data...")
            self._write_reconciled_data(merged_df, writer)
            
            # Sheet 2: Exceptions Only
            logger.info("Writing Sheet 2: Exceptions Only...")
            self._write_exceptions_sheet(merged_df, writer)
            
            # Sheet 3: Vendor Summary
            logger.info("Writing Sheet 3: Vendor Summary...")
            self._write_vendor_summary(merged_df, writer)
            
            # Sheet 4: YTD Summary
            logger.info("Writing Sheet 4: YTD Summary...")
            self._write_ytd_summary(merged_df, writer)
        
        # Format the Excel file
        logger.info("Formatting Excel file...")
        self._format_excel(filepath)
        
        logger.info(f"✓ Report generated: {filepath}")
        
        return filepath
    
    def _write_reconciled_data(self, df: pd.DataFrame, writer):
        """Write Sheet 1: Full reconciled data"""
        
        # Select columns for output
        columns = [
            "applicant_number", "consultant_name", "vendor_name", "service_month",
            "hours_billed", "hours_invoiced", "hours_paid",
            "bill_rate", "pay_rate",
            "amount_billed", "amount_invoiced", "amount_paid",
            "accrual_amount", "variance_hours", "variance_amount", "variance_pct",
            "exception_type", "exception_severity", "requires_action"
        ]
        
        # Filter to available columns
        available_cols = [col for col in columns if col in df.columns]
        output_df = df[available_cols].copy()
        
        # Sort by variance (descending)
        if "variance_amount" in output_df.columns:
            output_df = output_df.sort_values("variance_amount", ascending=False)
        
        output_df.to_excel(writer, sheet_name="Reconciled Data", index=False)
        
        logger.info(f"  Wrote {len(output_df)} records to Reconciled Data sheet")
    
    def _write_exceptions_sheet(self, df: pd.DataFrame, writer):
        """Write Sheet 2: Exceptions only"""
        
        # Filter to exceptions
        exceptions = df[df.get("requires_action", False) == True].copy()
        
        # Select columns
        columns = [
            "applicant_number", "consultant_name", "vendor_name",
            "hours_billed", "hours_paid", "amount_billed", "amount_paid",
            "accrual_amount", "variance_amount",
            "exception_type", "exception_severity", "exception_description"
        ]
        
        available_cols = [col for col in columns if col in exceptions.columns]
        output_df = exceptions[available_cols].copy()
        
        # Sort by severity (high, medium, low)
        if "exception_severity" in output_df.columns:
            severity_order = {"high": 0, "medium": 1, "low": 2}
            output_df["severity_rank"] = output_df["exception_severity"].map(severity_order)
            output_df = output_df.sort_values("severity_rank").drop("severity_rank", axis=1)
        
        output_df.to_excel(writer, sheet_name="Exceptions Only", index=False)
        
        logger.info(f"  Wrote {len(output_df)} exception records to Exceptions sheet")
    
    def _write_vendor_summary(self, df: pd.DataFrame, writer):
        """Write Sheet 3: Vendor summary (aggregated)"""
        
        # Group by vendor and service month
        group_cols = ["vendor_name"]
        if "service_month" in df.columns:
            group_cols.append("service_month")
        
        summary = df.groupby(group_cols).agg({
            "applicant_number": "count",
            "hours_billed": "sum",
            "amount_billed": "sum",
            "amount_paid": "sum",
            "accrual_amount": "sum",
            "requires_action": lambda x: (x == True).sum() if len(x) > 0 else 0
        }).reset_index()
        
        # Rename columns for clarity
        summary = summary.rename(columns={
            "applicant_number": "consultant_count",
            "requires_action": "exception_count"
        })
        
        # Calculate percentages
        summary["accrual_pct"] = (
            (summary["accrual_amount"] / (summary["amount_billed"] + 1) * 100).round(2)
        )
        
        # Sort by variance (descending)
        if "accrual_amount" in summary.columns:
            summary = summary.sort_values("accrual_amount", ascending=False)
        
        summary.to_excel(writer, sheet_name="Vendor Summary", index=False)
        
        logger.info(f"  Wrote {len(summary)} vendor summary records")
    
    def _write_ytd_summary(self, df: pd.DataFrame, writer):
        """Write Sheet 4: YTD summary (tie-back)"""
        
        # Get YTD columns
        ytd_cols = [
            "applicant_number", "consultant_name", "vendor_name",
            "ytd_hours_billed", "ytd_hours_paid",
            "ytd_amount_billed", "ytd_amount_paid", "ytd_accrual_amount"
        ]
        
        # Filter to available columns
        available_cols = [col for col in ytd_cols if col in df.columns]
        
        if available_cols:
            # Drop duplicates by applicant
            output_df = df[available_cols].drop_duplicates(subset=["applicant_number"])
            
            # Sort by YTD accrual (descending)
            if "ytd_accrual_amount" in output_df.columns:
                output_df = output_df.sort_values("ytd_accrual_amount", ascending=False)
            
            output_df.to_excel(writer, sheet_name="YTD Summary", index=False)
            
            logger.info(f"  Wrote {len(output_df)} YTD summary records")
    
    def _format_excel(self, filepath: Path):
        """Format Excel file with colors, fonts, column widths"""
        logger.info("Formatting Excel file...")
        
        try:
            workbook = load_workbook(filepath)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Set column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Freeze header row
                worksheet.freeze_panes = "A2"
            
            workbook.save(filepath)
            logger.info("✓ Excel file formatted successfully")
        
        except Exception as e:
            logger.warning(f"Formatting skipped (file still valid): {str(e)}")


if __name__ == "__main__":
    from loguru import logger
    logger.add("logs/generate_report.log")
    
    print("✓ ReportGenerator module ready")
